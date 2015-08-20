from openpyxl import load_workbook
from openpyxl import Workbook

class functions:
	
	def header(self):
		
		ws['A1'] = "Day"
		ws['B1'] = "Date"
		ws['C1'] = "Venue"
		ws['D1'] = "E"
		ws['E1'] = "Status"
		ws['F1'] = "Time"
		ws['G1'] = "Event"
		ws['H1'] = "FS Initials"
		ws['I1'] = "Comments"

def columnWidth(writer, dimensions):
	for d in dimensions:
		writer.column_dimensions[d[0]].width = d[1]
	return writer
	#writer.column_dimensions['B'].width = 20
	#ws.column_dimensions['G'].width = 40
	#ws.column_dimensions['I'].width = 40
